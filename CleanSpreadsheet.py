import sys,getopt
from openpyxl import load_workbook
from openpyxl import Workbook

def insertrow(worksheet,rownumber,rowcontents):
	for colnumber in range(1,len(rowcontents)+1):
		newcell = worksheet.cell(row=rownumber,column=colnumber)
		newcell.value=rowcontents[colnumber-1]
	
opts, args = getopt.getopt(sys.argv[1:],"i:o:")
inputfile = ''
outputfile = ''

for opt, arg in opts:
      if opt == '-h':
         print ('CleanSpreadsheet.py -i <inputfile> -o <outputfile>')
         sys.exit()
      elif opt in("-i", "--ifile"):
         inputfile = arg
      elif opt in ("-o", "--ofile"):
         outputfile = arg
print (inputfile)
print (outputfile)

wb = load_workbook(inputfile)
newworkbook=Workbook()
newworksheet=newworkbook.active
colheaders=[]
newrownumber=0
for ws in wb:
	rowcount=ws.max_row 
	colcount=ws.max_column 
	skiprow=False
	for r in range (1,rowcount +1):
		print("rownumber:",r)
		newrow=[]
		if ws.cell(row=r,column=colcount).value is not None:
			for c in range(1,colcount+1):
				newcell = ws.cell(row=r, column=c)
				if newcell.value is not None:
					newrow.append(newcell.value)
				
			if len(newrow)==colcount:
				skiprow=False
				if len(colheaders)!=colcount:
					colheaders=newrow
				elif colheaders==newrow:
						skiprow=True
				if not skiprow:
					newrownumber+=1
					insertrow(newworksheet,newrownumber,newrow)
							
newworkbook.save(outputfile)


								
